def polygon_to_bbox(polygon):
    # Convert polygon to bounding box
    x_coords, y_coords = zip(*polygon)
    min_x, max_x = min(x_coords), max(x_coords)
    min_y, max_y = min(y_coords), max(y_coords)
    w = round(max_x - min_x, 7)
    h = round(max_y - min_y, 7)
    return min_x, min_y, max_x, max_y, w, h

def pre_process_projection_label_file(projection_label_file_path):
    with open(projection_label_file_path, 'r') as infile:
        lines = infile.read().splitlines()

    bounding_boxes_data = []
    i = 0
    for line in lines:
        try:
            data = line.split()
            category = data[0]
            polygon = []
            for i in range(1, len(data), 2):
                x = float(data[i])
                y = float(data[i + 1])
                polygon.append((x, y))
            bbox = tuple(polygon_to_bbox(polygon))
            i = i + 1

            bounding_boxes_data.append(bbox)
        except:
#             print(projection_label_file_path + str(i) + "行出现问题！" + str(lines[i]))
            i = i + 1
            continue

    return bounding_boxes_data

def pre_process_ear_label_file(ear_file_path):
    polygon = []
    with open(ear_file_path, 'r') as file:
        data = file.readline().split()
        category = data[0]
        for i in range(1, len(data), 2):
            x = float(data[i])
            y = float(data[i + 1])
            polygon.append((x, y))

    polygon = np.array(polygon, dtype=np.float32)

    return polygon


from shapely.geometry import LineString, Polygon
import math

# 设定穗行数计算一圈的起始点
start_index_horizontal_row_count, end_index_horizontal_row_count = 0.055, 0.945

def calculate_kernel_row_number(bboxes_data):
    # 筛选出处于横向边界内的多边形
    bboxes_data = [bbox for bbox in bboxes_data if
                   start_index_horizontal_row_count < (bbox[0] + bbox[2]) / 2 < end_index_horizontal_row_count]

    # 排序后取中间1/3的bbox进行操作
    bboxes_data.sort(key=lambda x: x[3])  # 按 y_max 进行排序

    start_index = int(len(bboxes_data) / 3)
    end_index = int(len(bboxes_data) * 2 / 3)

    bboxes_data = bboxes_data[start_index:end_index]

    # 因为识别对象不分具体的籽粒种类，因此删除bbox数据中的种类标识
    ymin_values = [bbox[1] for bbox in bboxes_data]
    ymax_values = [bbox[3] for bbox in bboxes_data]

    LOW = min(ymin_values)
    HIGH = max(ymax_values)

    num_intervals = 1000
    step_size = (HIGH - LOW) / num_intervals

    # 将籽粒所在区间纵向等分，计算边界线穿过的bbox数量
    row_counts = []
    threshold = LOW

    for i in range(num_intervals):
        threshold = LOW + i * step_size
        row_bbox_count = sum(1 for bbox in bboxes_data if bbox[1] < threshold < bbox[3])

        raw_row_count = (row_bbox_count + 1) // 2 * 2  # 穗行数，取整（1r）
        row_counts.append(raw_row_count)

    # 选择出现频率最高的穗行数作为最终的穗行数
    kernel_row_number = max(set(row_counts), key=row_counts.count)

    return kernel_row_number


# 设定穗粒数计算一圈的起始点
start_index_horizontal_kernel_number, end_index_horizontal_kernel_number = 0.04, 0.96


def calculate_kernel_number(bboxes_data):
    # 筛选出处于横向边界内的多边形
    bboxes_data = [bbox for bbox in bboxes_data if
                   start_index_horizontal_kernel_number < (bbox[0] + bbox[2]) / 2 < end_index_horizontal_kernel_number]

    # 计算处于一圈界限内的多边形个数作为
    kernel_number = len(bboxes_data)

    return kernel_number


# 自定义粒厚计算公式
def calculate_kernel_temp_value(bboxes_data):
    # 筛选出处于横向边界内的多边形
    bboxes_data = [bbox for bbox in bboxes_data if
                   start_index_horizontal_kernel_number < (bbox[0] + bbox[2]) / 2 < end_index_horizontal_kernel_number]

    # 排序后取中间1/3的bbox进行操作
    bboxes_data.sort(key=lambda x: x[3])  # 按 y_max 进行排序

    start_index = int(len(bboxes_data) / 3)
    end_index = int(len(bboxes_data) * 2 / 3)

    bboxes_data = bboxes_data[start_index:end_index]

    ymin_values = [bbox[1] for bbox in bboxes_data]
    ymax_values = [bbox[3] for bbox in bboxes_data]

    LOW = min(ymin_values)
    HIGH = max(ymax_values)

    kernels_area_height = HIGH - LOW
    inside_kernel_number = len(bboxes_data)

    kernel_temp_value = kernels_area_height / inside_kernel_number

    return kernel_temp_value


def calculate_ear_length_width(polygon):
    # 直接引用计算
    (cx, cy), (l, w), theta = cv2.minAreaRect(polygon)  # 求出最小外接矩形，更严谨，避免了棒子歪斜带来的误差

    polygon = Polygon(polygon)  # 转换为多边形类型

    cutting_line = create_cutting_line(cx, cy, theta, l, w)

    # 获取线段与多边形边界的交点
    intersection_line = cutting_line.intersection(polygon)
    # ear_length = 0.2585716 + 0.01987238 * 1440 * max(l, w)  # 穗长的拟合曲线

    ear_length = max(l,w) * height_scale_ratio
    ear_width = intersection_line.length * width_scale_ratio

    return ear_length, ear_width


def calculate_ear_volume(polygon):
    distances = find_intersection_distances(polygon)
    ear_volume = 0
    for distance in distances:
        radius = distance / 2 * width_scale_ratio
        h = 0.01 * height_scale_ratio
        ear_volume += math.pi * radius ** 2 * h

    return ear_volume


def create_cutting_line(cx, cy, theta, l, w):
    """
    根据中心点、旋转角度和长短边创建一条垂直于多边形长边的线段。
    """
    if l >= w:
        # 如果长边是l，计算垂直于长边的线段
        rad = np.deg2rad(-theta)
        dx = np.sin(rad) * 1000  # 1000是线段长度的一半，可以根据需要调整
        dy = np.cos(rad) * 1000
    else:
        #     如果长边是w，计算垂直于短边（也就是长边）的线段
        rad = np.deg2rad(theta)  # 加90度使线段垂直于长边
        dx = np.cos(rad) * 1000
        dy = np.sin(rad) * 1000

    line = LineString([(cx - dx, cy - dy), (cx + dx, cy + dy)])
    return line


def find_intersection_distances(polygon):
    # 获取线段与多边形相交部分的长度列表
    num_lines = 100
    intersection_distances = []
    polygon = Polygon(polygon)  # 将array转化为geometry形式才能操作

    for i in range(1, num_lines):
        y = i / 100
        line = LineString([(0, y), (1, y)])

        # 获取线段与多边形边界的交点
        intersection_line = line.intersection(polygon)

        if intersection_line.is_empty:
            continue

        if intersection_line.geom_type == 'LineString':
            intersection_distance = intersection_line.length
            intersection_distances.append(intersection_distance)
        elif intersection_line.geom_type == 'MultiLineString':
            # 处理 MultiLineString，获取内部各个 LineString 的长度
            for part in intersection_line.geoms:
                intersection_distance = part.length
                intersection_distances.append(intersection_distance)

    return intersection_distances


def average(data_list):
    if len(data_list) == 0:
        return None
    return sum(data_list) / len(data_list)


# scale_ratio为转换系数，最终计算单位为cm(2cm-100pixels, image_height=1440, image_width=1072)
height_scale_ratio = 2/100*1440
width_scale_ratio = 2/100*1072

def process_projection_file(file_path):
    # 预处理投影预测文件获取bboxes数据
    bounding_boxes_data = pre_process_projection_label_file(file_path)

    # 穗粒数
    kernel_number = calculate_kernel_number(bounding_boxes_data)

    # 穗行数
    kernel_row_number = calculate_kernel_row_number(bounding_boxes_data)
    kernel_row_number = round(kernel_row_number, 0)

    # 行粒数
    try:
        kernel_number_per_row = kernel_number // kernel_row_number
    except:
        kernel_number_per_row = "NA"

    # 粒厚
    kernel_temp_value = calculate_kernel_temp_value(bounding_boxes_data)
    kernel_thickness = kernel_temp_value * kernel_row_number * height_scale_ratio

    return kernel_number, kernel_row_number, kernel_number_per_row, kernel_thickness


def process_ear_file(file_path):
    # 预处理整穗预测文件
    polygon = pre_process_ear_label_file(file_path)

    # 穗长穗宽
    ear_length, ear_width = calculate_ear_length_width(polygon)

    # 穗体积估算
    ear_volume = calculate_ear_volume(polygon)

    return ear_length, ear_width, ear_volume

from tensorflow.keras.models import load_model

# 定义穗行分类图像处理函数
def preprocess_image(image_path, target_size):
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)  # 使用OpenCV读取图像
    image = cv2.resize(image, target_size)  # 调整图像尺寸
    image = np.expand_dims(image, axis=-1)  # 增加一个通道维度
    image = np.expand_dims(image, axis=0)  # 增加一个批次维度
    return image

def kernel_row_number_classification(image_path, model):
    image = preprocess_image(image_path, (640, 640))
    prediction = np.argmax(model.predict(image)) # 返回预测类别
    return prediction

def phenotype_classification(image_path, model):
    image = preprocess_image(image_path, (640, 640))
    prediction = np.argmax(model.predict(image)) # 返回预测类别
    return prediction


from openpyxl import load_workbook, Workbook
import glob
import os
import numpy as np
import cv2
import argparse

if __name__ == "__main__":
    # Create the parser
    parser = argparse.ArgumentParser(description="Analyze ear and projection model output data")

    # Add arguments
    parser.add_argument('-i', '--projection_image_folder', default='./images/projection/', type=str, required=True, help='Path to the projection image folder')
    parser.add_argument('-e', '--ear_label_folder', default='./result/ear/labels/', type=str, required=True, help='Path to the ear label folder')
    parser.add_argument('-p', '--projection_label_folder', default='./result/projection/labels/', type=str, required=True, help='Path to the projection label folder')
    parser.add_argument('-o', '--output_path', default='./', type=str, required=True, help='Output file path')
    parser.add_argument('-m', '--model_path', default='./models/', type=str, required=True, help='CNN model path')

    # Parse the arguments
    args = parser.parse_args()

    projection_label_files = sorted(glob.glob(os.path.join(args.projection_label_folder, '*.txt')))
    output_file = os.path.join(args.output_path, "results.xlsx")

    wb = Workbook()

    ws = wb.active
    ws.append(
        ["Labels", "Predicted_Ear_Length", "Predicted_Ear_Width", "Predicted_Ear_Volume", "Predicted_Kernel_Number",
         "Predicted_Kernel_Row_Number", "Predicted_Kernel_Number_per_Row", "Predicted_Kernel_Thickness"])

    for projection_label_file in projection_label_files:
        label = projection_label_file.split("/")[-1].split(".")[0]  # 提取文件名
        kernel_number, kernel_row_number, kernel_number_per_row, kernel_thickness = process_projection_file(projection_label_file)

        ear_label_files = glob.glob(os.path.join(args.ear_label_folder, label + '_phenotyping_?.txt'))
        ear_lengths, ear_widths, ear_volumes = [], [], []
        for ear_label_file in ear_label_files:
            ear_length, ear_width, ear_volume = process_ear_file(ear_label_file)
            ear_lengths.append(ear_length)
            ear_widths.append(ear_width)
            ear_volumes.append(ear_volume)

        ear_length = round(np.median(np.array(ear_lengths)), 2)
        ear_width = round(np.median(np.array(ear_widths)), 2)
        ear_volume = round(np.median(np.array(ear_volumes)), 2)

        kernel_thickness = round(kernel_thickness, 2)

        # 添加数据到Excel表格
        ws.append([label, ear_length, ear_width, ear_volume, kernel_number, kernel_row_number, kernel_number_per_row,
                   kernel_thickness])

    wb.save(output_file)

    # 加载模型
    model1 = load_model(os.path.join(args.model_path, "1_Developmental_Status_Assesment.h5"))
    model2 = load_model(os.path.join(args.model_path, "2_Kernel_Row_Visibility_Assesment.h5"))
    wb = load_workbook(output_file)
    ws = wb.active

    pic_path = args.projection_image_folder

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # 假设标签在第一列
        label = row[0].value
        projection_pic_path = os.path.join(pic_path, f'{label}.png')
        phenotype_prediction = phenotype_classification(projection_pic_path, model1)
        kernel_row_number_prediction = kernel_row_number_classification(projection_pic_path, model2)
        # 根据model1的预测结果决定是否标记整行为NA
        if phenotype_prediction == 0:
            for col in range(2, ws.max_column + 1):  # 假设要标记整行
                ws.cell(row=row[0].row, column=col, value='NA')
        else:
            # 根据model2的预测结果决定是否只在第六列标记为NA
            if kernel_row_number_prediction == 0:
                ws.cell(row=row[0].row, column=6, value='NA')  # 假设kernel_row_number在第六列

    wb.save(output_file)